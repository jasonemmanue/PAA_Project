"""Import des données comparatives pluriannuelles — feuille SYNTHESE COMPAREE.

Fichier source : ALLER-RETOUR_ED_TRAITEMENT_DES_DONNEES_FEVRIER_2026.xlsx
Feuille        : 'SYNTHESE COMPAREE'
Table cible    : evolution_indicateur

Structure de la feuille (blocs de lignes) :
  L6  : en-têtes axes (CARENA Aller, Toyota Aller, SODECI Aller | retour × 3)
  L7  : noms complets des axes
  L9  : dates des périodes (datetime objects → 'oct_2025', 'fev_2026')
  L10 : temps minimal  — Jours ouvrables
  L11 : temps minimal  — Week-ends
  L12 : temps moyen    — Jours ouvrables
  L13 : temps moyen    — Week-ends
  L14 : temps maximal  — Jours ouvrables
  L15 : temps maximal  — Week-ends

Les colonnes sont en paires (une par période) pour chaque axe.
Toutes les durées sont en minutes dans la source ; converties en secondes ici.

Utilisation directe :
    python -m app.import_evolution chemin/vers/fichier.xlsx

Appelé aussi via POST /import/evolution.
"""

from __future__ import annotations

import logging
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db.session import SessionLocal
from app.models.models import EvolutionIndicateur

logger = logging.getLogger("paa.import.evolution")

FEUILLE = "SYNTHESE COMPAREE"

# Layout des colonnes (index 0-based) dans la feuille
# Chaque axe occupe 2 colonnes consécutives : (période_1, période_2)
_AXES_COLONNES: list[tuple[str, str, int, int]] = [
    # (axe_nom, sens, col_periode_1, col_periode_2)
    ("CARENA → Pharmacie Palm Beach",       "Aller",  3,  4),
    ("Toyota CFAO → Pharmacie Palm Beach",  "Aller",  5,  6),
    ("Agence SODECI → Pharmacie Palm Beach","Aller",  7,  8),
    ("Pharmacie Palm Beach → CARENA",       "Retour", 10, 11),
    ("Pharmacie Palm Beach → Toyota CFAO",  "Retour", 12, 13),
    ("Pharmacie Palm Beach → Agence SODECI","Retour", 14, 15),
]

# Ligne de départ du bloc de données (index 0-based, L9 → index 8)
_LIGNE_DATES = 8       # L9 dans Excel (0-based)
_LIGNE_DEBUT_METRIQUES = 9   # L10
_METRIQUES = [
    # (offset_depuis_LIGNE_DEBUT, type_jour, metrique)
    (0, "Jours ouvrables", "min"),
    (1, "Week-ends",       "min"),
    (2, "Jours ouvrables", "moyen"),
    (3, "Week-ends",       "moyen"),
    (4, "Jours ouvrables", "max"),
    (5, "Week-ends",       "max"),
]


def _format_periode(val) -> str | None:
    """Convertit une valeur de cellule date en code de période ('oct_2025')."""
    if val is None:
        return None
    if isinstance(val, datetime):
        mois = {1:"jan",2:"fev",3:"mar",4:"avr",5:"mai",6:"jun",
                7:"jul",8:"aou",9:"sep",10:"oct",11:"nov",12:"dec"}
        return f"{mois[val.month]}_{val.year}"
    # Parfois openpyxl retourne un string si data_only ne résout pas la formule
    try:
        dt = datetime.fromisoformat(str(val))
        return _format_periode(dt)
    except (ValueError, TypeError):
        return str(val).strip()


def _min_vers_s(val) -> float | None:
    """Convertit des minutes (float) en secondes (float), None si absent."""
    if val is None:
        return None
    try:
        return float(val) * 60.0
    except (ValueError, TypeError):
        return None


def importer(chemin_fichier: str | Path) -> dict:
    """Importe la feuille SYNTHESE COMPAREE dans evolution_indicateur.

    Retourne un dict résumant les résultats.
    """
    chemin = Path(chemin_fichier)
    logger.info("Lecture de %s (feuille '%s')…", chemin.name, FEUILLE)

    wb = openpyxl.load_workbook(str(chemin), read_only=True, data_only=True)
    if FEUILLE not in wb.sheetnames:
        raise ValueError(
            f"Feuille '{FEUILLE}' introuvable dans {chemin.name}. "
            f"Feuilles disponibles : {wb.sheetnames}"
        )
    ws = wb[FEUILLE]
    toutes_lignes = list(ws.iter_rows(values_only=True))
    wb.close()

    # Lecture des périodes depuis la ligne de dates
    ligne_dates = toutes_lignes[_LIGNE_DATES]
    periodes: dict[int, str | None] = {}
    for _, _, col1, col2 in _AXES_COLONNES:
        if periodes.get(col1) is None:
            periodes[col1] = _format_periode(ligne_dates[col1])
        if periodes.get(col2) is None:
            periodes[col2] = _format_periode(ligne_dates[col2])

    logger.info("Périodes détectées : %s", dict(sorted(periodes.items())))

    # Construction des enregistrements à insérer
    enregistrements: list[dict] = []
    for offset, type_jour, metrique in _METRIQUES:
        ligne = toutes_lignes[_LIGNE_DEBUT_METRIQUES + offset]
        for axe, sens, col1, col2 in _AXES_COLONNES:
            for col in (col1, col2):
                periode = periodes.get(col)
                if not periode:
                    continue
                valeur_s = _min_vers_s(ligne[col])
                enregistrements.append({
                    "axe": axe,
                    "sens": sens,
                    "periode": periode,
                    "type_jour": type_jour,
                    "metrique": metrique,
                    "valeur_s": valeur_s,
                })

    # Regrouper par (axe, sens, periode, type_jour) pour construire les objets
    groupes: dict[tuple, dict] = {}
    for rec in enregistrements:
        cle = (rec["axe"], rec["sens"], rec["periode"], rec["type_jour"])
        if cle not in groupes:
            groupes[cle] = {"temps_min_s": None, "temps_moyen_s": None, "temps_max_s": None}
        if rec["metrique"] == "min":
            groupes[cle]["temps_min_s"] = rec["valeur_s"]
        elif rec["metrique"] == "moyen":
            groupes[cle]["temps_moyen_s"] = rec["valeur_s"]
        elif rec["metrique"] == "max":
            groupes[cle]["temps_max_s"] = rec["valeur_s"]

    logger.info("%d enregistrements à insérer.", len(groupes))

    session = SessionLocal()
    nb_inseres = 0
    nb_ignores = 0
    nb_erreurs = 0

    try:
        for (axe, sens, periode, type_jour), stats in groupes.items():
            # Idempotence : on vérifie avant d'insérer
            existant = (
                session.query(EvolutionIndicateur)
                .filter_by(axe=axe, sens=sens, periode=periode, type_jour=type_jour)
                .first()
            )
            if existant:
                nb_ignores += 1
                continue

            # On skip les lignes où toutes les valeurs sont None
            if all(v is None for v in stats.values()):
                logger.debug("Valeurs toutes NULL pour %s %s %s %s — ignoré.",
                             axe, sens, periode, type_jour)
                nb_ignores += 1
                continue

            try:
                obj = EvolutionIndicateur(
                    axe=axe,
                    sens=sens,
                    periode=periode,
                    type_jour=type_jour,
                    **stats,
                )
                session.add(obj)
                nb_inseres += 1
            except Exception as exc:
                logger.warning("Erreur insertion %s/%s/%s/%s : %s",
                               axe, sens, periode, type_jour, exc)
                nb_erreurs += 1

        session.commit()

    except Exception:
        session.rollback()
        raise
    finally:
        session.close()

    resultat = {
        "nb_groupes_detectes": len(groupes),
        "nb_inseres": nb_inseres,
        "nb_ignores_doublon": nb_ignores,
        "nb_erreurs": nb_erreurs,
        "periodes": list({p for p in periodes.values() if p}),
    }
    logger.info("Import terminé : %s", resultat)
    _afficher_resume(resultat)
    return resultat


def _afficher_resume(r: dict) -> None:
    print(f"\n{'=' * 55}")
    print(f"  Import SYNTHESE COMPAREE — résumé")
    print(f"{'=' * 55}")
    print(f"  Groupes détectés     : {r['nb_groupes_detectes']}")
    print(f"  Insérés              : {r['nb_inseres']}")
    print(f"  Ignorés (doublon)    : {r['nb_ignores_doublon']}")
    print(f"  Erreurs              : {r['nb_erreurs']}")
    print(f"  Périodes             : {', '.join(r['periodes'])}")
    print(f"{'=' * 55}\n")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    if len(sys.argv) < 2:
        print("Usage : python -m app.import_evolution <chemin_fichier.xlsx>")
        sys.exit(1)
    importer(sys.argv[1])
